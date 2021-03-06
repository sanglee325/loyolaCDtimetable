﻿# -*- coding: utf-8 -*-
import openpyxl
from loyolaCD import employee as E

def get_info(filename, new, wb, ws):
        #save data from filename
        tmp = filename.split('_')
        tmp2 = tmp[2].split('.')

        new.student_num = tmp[1]
        new.name = tmp2[0]
        new.major = ws['E22'].value
        new.phone_num = ws['G22'].value
        new.late = ws['I39'].value


def read_timetable(new, wb, ws):
        #get week timetable
        for i in range(9):
                for j in range(5):
                        info = ws.cell(row=26+i, column=5+j).value
                        if info == None:
                                new.week[i][j] = '-'
                        else:
                                tmp = list(info)
                                if tmp[0] == 'O' or tmp[0] == 'o':
                                        new.week[i][j] = 'O'
                                elif tmp[0] == 'X' or tmp[0] == 'x':
                                        new.week[i][j] = 'X'
                        '''
                        if info == None:
                                new.week[i][j] = '-'
                        else:
                                tmp = list(info)
                                new.week[i][j] = tmp[0]
                        '''
        #get weekend timetable
        for j in range(2):
                info = ws.cell(row=38+j, column=5).value
                if info == None:
                        new.weekend[j] = '-'
                else:   
                        tmp = list(info)
                        if tmp[0] == 'O' or tmp[0] == 'o':
                                new.weekend[j] = 'O'
                        elif tmp[0] == 'X' or tmp[0] == 'x':
                                new.weekend[j] = 'X'

def divided_week(week, student):
        tmp1 = week.split('[')
        tmp2 = "".join(tmp1)
        tmp1 = tmp2.split(']')
        tmp2 = "".join(tmp1)
        tmp1 = tmp2.split(',')
        tmp2 = "".join(tmp1)
        tmp1 = tmp2.split("'")
        tmp2 = "".join(tmp1)
        tmp1 = tmp2.split(" ")
        tmp2 = "".join(tmp1)
        k = 0
        for i in range(9):
                for j in range(5):
                        student[i][j] = tmp2[k]
                        k += 1

def divided_weekend(weekend, student):
        tmp1 = weekend.split('[')
        tmp2 = "".join(tmp1)
        tmp1 = tmp2.split(']')
        tmp2 = "".join(tmp1)
        tmp1 = tmp2.split(',')
        tmp2 = "".join(tmp1)
        tmp1 = tmp2.split("'")
        tmp2 = "".join(tmp1)
        tmp1 = tmp2.split(" ")
        tmp2 = "".join(tmp1)

        student[0] = tmp2[0]
        student[1] = tmp2[1]
